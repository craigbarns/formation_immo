#!/usr/bin/env python3
"""Génère tous les fichiers Excel (.xlsx) des modules — exécuter depuis la racine formation-immobiliere."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
YELLOW = PatternFill("solid", fgColor="FFF2CC")
HEADER = PatternFill("solid", fgColor="1A3A5C")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def style_header(ws, row=1, cols=10):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def m1_mandats():
    path = ROOT / "module1-juridique/templates/TABLEAU_MANDATS_COMPARATIF.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparatif mandats"
    headers = [
        "Critère",
        "Mandat simple",
        "Mandat exclusif",
        "Mandat semi-exclusif",
        "Mandat co-exclusif",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))
    rows = [
        ("Nombre d'agences mandatées", "Illimité", "1 seule", "1 + possibilité mandat simple ailleurs", "2 agences max"),
        ("Exclusivité pour l'agence", "Non", "Oui", "Partielle", "Partagée entre co-mandataires"),
        ("Engagement vendeur", "Faible", "Fort", "Moyen", "Moyen à fort"),
        ("Honoraires indicatifs", "Négociables", "Souvent plus élevés", "Intermédiaire", "Partagés"),
        ("Intérêt pour l'agence", "Concurrence forte", "Priorisation marketing", "Compromis", "Partage du risque"),
    ]
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    ws.column_dimensions["A"].width = 28
    for col in range(2, 6):
        ws.column_dimensions[get_column_letter(col)].width = 22
    wb.save(path)
    return path


def m2_estimation():
    path = ROOT / "module2-transaction/templates/GRILLE_ESTIMATION_COMPARATIVE.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimation comparative"
    ws["A1"] = "Bien à estimer — [ADRESSE]"
    ws["A2"] = "Surface habitable (m²)"
    ws["B2"] = 78
    ws["A3"] = "Type"
    ws["B3"] = "Appartement T4"
    ws["A5"] = "Comparables vendus"
    hdr = ["Réf", "Adresse", "Vente (€)", "Surface m²", "Prix/m²", "Ajustement %", "Prix/m² ajusté"]
    for i, h in enumerate(hdr, 1):
        ws.cell(row=6, column=i, value=h)
    style_header(ws, 6, len(hdr))
    data = [
        ("C1", "Lyon 6e — Brotteaux", 485000, 76, "=C7/D7", 0, "=E7*(1+F7)"),
        ("C2", "Lyon 6e — Masséna", 510000, 82, "=C8/D8", -0.02, "=E8*(1+F8)"),
        ("C3", "Lyon 6e — Jean Macé", 465000, 74, "=C9/D9", 0.03, "=E9*(1+F9)"),
    ]
    for r, row in enumerate(data, 7):
        for c, val in enumerate(row, 1):
            if isinstance(val, str) and val.startswith("="):
                ws.cell(row=r, column=c, value=val)
            else:
                ws.cell(row=r, column=c, value=val)
    ws["A11"] = "Moyenne prix/m² ajusté (€)"
    ws["B11"] = "=AVERAGE(G7:G9)"
    ws["A12"] = "Fourchette basse (€)"
    ws["B12"] = "=B11*B2*0.97"
    ws["A13"] = "Prix conseillé (€)"
    ws["B13"] = "=B11*B2"
    ws["A14"] = "Fourchette haute (€)"
    ws["B14"] = "=B11*B2*1.03"
    for addr in ["B2", "B11", "B12", "B13", "B14"]:
        ws[addr].fill = YELLOW
    wb.save(path)
    return path


def m2_crm():
    path = ROOT / "module2-transaction/templates/TEMPLATE_SUIVI_CLIENT.xlsx"
    wb = Workbook()
    # Feuille clients
    ws = wb.active
    ws.title = "Clients"
    hc = ["Code", "Nom", "Type", "Téléphone", "Email", "Budget max", "Statut", "Prochain rappel"]
    for i, h in enumerate(hc, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(hc))
    exemples = [
        ("CL-001", "MARTIN", "Acquéreur", "06 00 00 00 01", "m.martin@email.fr", 450000, "Visite prévue", "2026-04-15"),
        ("CL-002", "DURAND", "Vendeur", "06 00 00 00 02", "durand@email.fr", "-", "Mandat signé", "2026-04-20"),
    ]
    for r, row in enumerate(exemples, 2):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    # Pipeline
    wp = wb.create_sheet("Pipeline")
    hp = ["Réf mandat", "Adresse", "Prix affiché", "Probabilité %", "Honoraires HT"]
    for i, h in enumerate(hp, 1):
        wp.cell(row=1, column=i, value=h)
    style_header(wp, 1, len(hp))
    wp["A2"] = "M-2026-01"
    wp["B2"] = "45 rue de Sèze, Lyon 6e"
    wp["C2"] = 485000
    wp["D2"] = 70
    wp["E2"] = "=C2*0.05"
    # Rappels
    wr = wb.create_sheet("Rappels")
    wr["A1"] = "Date"
    wr["B1"] = "Action"
    wr["C1"] = "Priorité"
    style_header(wr, 1, 3)
    wr.append(["2026-04-10", "Relance MARTIN", "Haute"])
    wb.save(path)
    return path


def m3_credit():
    path = ROOT / "module3-financement/templates/SIMULATEUR_CREDIT.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Simulateur"
    ws["A1"] = "SIMULATEUR CRÉDIT (indicatif — valider avec un courtier)"
    ws["A1"].font = Font(bold=True, size=14)
    # B3 prix, B4 apport, B5 emprunt = prix - apport, B6 durée ans, B7 taux annuel, B8 taux assurance/an sur capital
    pairs = [
        ("Prix du bien (€)", 350000),
        ("Apport (€)", 35000),
        ("Montant emprunté (€)", "=B3-B4"),
        ("Durée (années)", 20),
        ("Taux nominal annuel (ex: 0,035)", 0.035),
        ("Taux assurance / an sur capital (ex: 0,0034)", 0.0034),
    ]
    r = 3
    for lab, val in pairs:
        ws.cell(row=r, column=1, value=lab)
        ws.cell(row=r, column=2, value=val)
        r += 1
    ws["A9"] = "Mensualité hors assurance (€)"
    ws["B9"] = "=PMT(B7/12,B6*12,-B5)"
    ws["A10"] = "Mensualité assurance (approx., €)"
    ws["B10"] = "=B5*B8/12"
    ws["A11"] = "Mensualité totale (€)"
    ws["B11"] = "=B9+B10"
    ws["A13"] = "Revenus nets mensuels (€)"
    ws["B13"] = 5500
    ws["A14"] = "Autres crédits mensuels (€)"
    ws["B14"] = 200
    ws["A15"] = "Taux d'endettement (sur revenus)"
    ws["B15"] = "=B11/B13"
    ws["A16"] = "Reste à vivre estimé (€)"
    ws["B16"] = "=B13-B14-B11"
    for row in [3, 4, 5, 6, 7, 8, 13, 14]:
        ws.cell(row=row, column=2).fill = YELLOW
    wb.save(path)
    return path


def m3_fiscalite():
    path = ROOT / "module3-financement/templates/TABLEAU_FISCALITE.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Barèmes"
    ws["A1"] = "Indicateurs fiscaux — à actualiser selon loi de finances"
    ws["A3"] = "Pinel — réduction max (€) zone A bis / A / B1 (exemple)"
    ws["B3"] = 52500
    ws["A4"] = "Durée d'engagement (ans)"
    ws["B4"] = 12
    ws["A6"] = "IFI — tranche (€)"
    ws["B6"] = "Taux applicable"
    rows = [("0 — 800 000", "0%"), ("800 001 — 1 300 000", "0,5%"), ("1 300 001 — 2 570 000", "0,7%")]
    r = 7
    for a, b in rows:
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        r += 1
    wb.save(path)
    return path


def m3_renta():
    path = ROOT / "module3-financement/templates/CALCUL_RENTABILITE.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Rentabilité"
    ws["A1"] = "CALCUL RENTABILITÉ LOCATIF (indicatif)"
    data = [
        ("Prix d'achat HT", 180000),
        ("Travaux", 15000),
        ("Meubles", 5000),
        ("Loyer mensuel HC", 750),
        ("Charges annuelles (taxe foncière, copro...)", 2400),
        ("Mois de vacance locative / an", 1),
    ]
    r = 3
    for lab, val in data:
        ws.cell(row=r, column=1, value=lab)
        ws.cell(row=r, column=2, value=val)
        ws.cell(row=r, column=2).fill = YELLOW
        r += 1
    ws["A10"] = "Loyer annuel ajusté vacance"
    ws["B10"] = "=B6*(12-B9)"
    ws["A11"] = "Rendement brut"
    ws["B11"] = "=B10/(B3+B4+B5)"
    ws["A12"] = "Rendement net de charges"
    ws["B12"] = "=(B10-B8)/(B3+B4+B5)"
    ws["A13"] = "Cash-flow mensuel après charges"
    ws["B13"] = "=(B10-B8)/12"
    wb.save(path)
    return path


def m4_calendrier():
    path = ROOT / "module4-marketing/templates/CALENDRIER_EDITORIAL_RESEAUX.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Calendrier"
    ws["A1"] = "Calendrier éditorial — [AGENCE] — Avril 2026"
    h = ["Date", "Réseau", "Format", "Thème", "Statut"]
    for i, x in enumerate(h, 1):
        ws.cell(row=3, column=i, value=x)
    style_header(ws, 3, 5)
    rows = [
        ("2026-04-08", "Instagram", "Reel", "Coulisses visite", "Prêt"),
        ("2026-04-10", "LinkedIn", "Post", "Analyse marché local", "Brouillon"),
        ("2026-04-12", "Facebook", "Photo", "Nouvelle exclusivité", "Planifié"),
    ]
    for r, row in enumerate(rows, 4):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    wb.save(path)
    return path


def m4_kpi():
    path = ROOT / "module4-marketing/templates/TABLEAU_PERFORMANCE_KPIs.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = "KPI Marketing — [MOIS/ANNÉE]"
    h = ["Indicateur", "Valeur", "Objectif", "Écart"]
    for i, x in enumerate(h, 1):
        ws.cell(row=3, column=i, value=x)
    style_header(ws, 3, 4)
    kpis = [
        ("Visites site web", 1250, 1500),
        ("Leads qualifiés", 42, 50),
        ("Coût par lead (€)", 35, 30),
        ("Taux conversion visite→lead %", 3.4, 4.0),
    ]
    r = 4
    for name, val, obj in kpis:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=val)
        ws.cell(row=r, column=3, value=obj)
        ws.cell(row=r, column=4, value=f"=C{r}-B{r}")
        r += 1
    wb.save(path)
    return path


def main():
    out = []
    for fn in (m1_mandats, m2_estimation, m2_crm, m3_credit, m3_fiscalite, m3_renta, m4_calendrier, m4_kpi):
        p = fn()
        out.append(str(p.relative_to(ROOT)))
    print("Fichiers générés :")
    for x in out:
        print(" ", x)


if __name__ == "__main__":
    main()
